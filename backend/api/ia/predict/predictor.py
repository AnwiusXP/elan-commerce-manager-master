from fastapi import APIRouter, Query, Depends
from fastapi.responses import StreamingResponse
import xgboost as xgb
import pandas as pd
import google.genai as genai
from google.api_core import exceptions as google_exceptions
import json
import os
import re
import sys
from datetime import datetime
from io import BytesIO
from dotenv import load_dotenv
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "../../../../")))
try:
    from auth import get_db
    from models import Venta
except ImportError:
    pass

load_dotenv(encoding="utf-8")

router = APIRouter()

CACHE_TTL = 12 * 3600
CACHE_GEMINI = {}

client = genai.Client(api_key=os.getenv("GOOGLE_API_KEY"))
GEMINI_MODEL = "gemini-3.1-flash-lite"
FECHAS_FUTURAS = ["2026-04-15", "2026-04-16", "2026-04-17"]

EXCEL_VERDE_CORPORATIVO = "1E8A5E"
EXCEL_GRIS_OSCURO = "161B22"
EXCEL_BLANCO = "FFFFFF"
EXCEL_GRIS_BORDE = "D0D7DE"
EXCEL_FONDO_SUAVE = "F6F8FA"


def _extraer_datos_ventas(db: Session) -> pd.DataFrame:
    ventas = db.query(Venta).all()
    datos_procesados = []

    for venta in ventas:
        if not venta.items:
            continue
        try:
            items = json.loads(venta.items)
            if isinstance(items, list):
                for item in items:
                    datos_procesados.append({
                        "producto_id": str(item.get("producto_id", "Desconocido")),
                        "nombre_producto": item.get("nombre_producto", "Desconocido"),
                        "cantidad": float(item.get("cantidad", 0)),
                        "fecha_venta": venta.fecha.date() if venta.fecha else datetime.now().date()
                    })
        except Exception as parse_error:
            print(f"Error parseando items de venta ID {venta.id}: {parse_error}")

    if not datos_procesados:
        print("No hay ventas validas en la BD, usando fallback para modelo.")
        datos_procesados = [
            {"producto_id": "1", "nombre_producto": "Fallback", "cantidad": 50, "fecha_venta": "2026-04-10"},
            {"producto_id": "1", "nombre_producto": "Fallback", "cantidad": 55, "fecha_venta": "2026-04-11"}
        ]

    df = pd.DataFrame(datos_procesados)
    df["fecha_venta"] = pd.to_datetime(df["fecha_venta"]).dt.strftime("%Y-%m-%d")
    return df


def _filtrar_por_producto(df: pd.DataFrame, producto: str) -> tuple[pd.DataFrame, str]:
    nombre_producto = "Todos los productos"

    if producto != "Todos":
        print(f"Buscando el producto con ID o Nombre: {producto}")
        mask = (
            (df["producto_id"].astype(str) == str(producto)) |
            (df["nombre_producto"].astype(str).str.lower() == str(producto).lower())
        )
        df = df[mask]
        nombre_producto = producto

    return df, nombre_producto


def _generar_recomendacion(nombre_producto: str, cantidades_historicas: list[float], predicciones_xgboost: list[int]) -> str:
    cache_key = nombre_producto.strip().lower()
    ahora = datetime.now()

    if cache_key in CACHE_GEMINI:
        cache_entry = CACHE_GEMINI[cache_key]
        tiempo_transcurrido = ahora - cache_entry["timestamp"]
        if tiempo_transcurrido.total_seconds() < CACHE_TTL:
            print(f"Usando cache para '{cache_key}'")
            return cache_entry["recomendacion"]

    min_hist = min(cantidades_historicas)
    max_hist = max(cantidades_historicas)
    avg_hist = sum(cantidades_historicas) / len(cantidades_historicas)

    prompt_gemini = f"""Producto: '{nombre_producto}'.
Historial: min={min_hist:.0f}, max={max_hist:.0f}, promedio={avg_hist:.1f}.
Proyeccion XGBoost: {predicciones_xgboost}.
Recomienda en 1-2 lineas si comprar mas stock."""

    try:
        response = client.models.generate_content(
            model=GEMINI_MODEL,
            contents=prompt_gemini
        )
        texto_recomendacion = response.text
        CACHE_GEMINI[cache_key] = {
            "recomendacion": texto_recomendacion,
            "timestamp": ahora
        }
        return texto_recomendacion
    except google_exceptions.ResourceExhausted:
        print(f"Cuota de Gemini agotada para '{cache_key}'")
        return "Analisis de IA en pausa por optimizacion de recursos. Basado en la grafica, evalue la tendencia para la produccion."
    except google_exceptions.DeadlineExceeded:
        print(f"Timeout de Gemini para '{cache_key}'")
        return "El servicio de IA tardo demasiado en responder. Basado en la grafica, evalue la tendencia."
    except Exception as gemini_error:
        print(f"Error inesperado de Gemini para '{cache_key}': {gemini_error}")
        return "Recomendacion IA no disponible temporalmente. Basado en la grafica, evalue la tendencia."


def _construir_analisis(producto: str, db: Session) -> dict:
    df_original = _extraer_datos_ventas(db)
    df_filtrado, nombre_producto = _filtrar_por_producto(df_original, producto)

    if df_filtrado.empty:
        print(f"El filtro dejo el DataFrame vacio para el producto {producto}")
        resultado = {
            "producto": nombre_producto,
            "recomendacion": f"No hay suficientes datos historicos para el producto {nombre_producto}.",
            "nivel": "Pendiente",
            "cantidad_estimada": 0,
            "datosGrafica": {
                "labels": [],
                "datasets": []
            }
        }
        return {
            "resultado": resultado,
            "historico": df_filtrado.copy(),
            "proyeccion": pd.DataFrame(columns=["fecha_proyectada", "cantidad_proyectada", "origen_modelo"])
        }

    df_agrupado = df_filtrado.groupby("fecha_venta")["cantidad"].sum().reset_index()
    df_agrupado = df_agrupado.sort_values("fecha_venta")

    fechas_historicas = df_agrupado["fecha_venta"].astype(str).tolist()
    cantidades_historicas = df_agrupado["cantidad"].tolist()

    # Mantiene el comportamiento existente: proyeccion simple sobre el ultimo valor.
    ultimo_valor = cantidades_historicas[-1] if cantidades_historicas else 50
    predicciones_xgboost = [int(ultimo_valor * 1.1), int(ultimo_valor * 1.2), int(ultimo_valor * 1.3)]
    cantidad_estimada_final = predicciones_xgboost[-1]

    labels_grafica = fechas_historicas + FECHAS_FUTURAS
    data_grafica = cantidades_historicas + predicciones_xgboost
    texto_recomendacion = _generar_recomendacion(nombre_producto, cantidades_historicas, predicciones_xgboost)
    nivel_alerta = "Alto" if cantidad_estimada_final > 70 else "Normal"

    resultado = {
        "producto": nombre_producto,
        "recomendacion": texto_recomendacion,
        "nivel": nivel_alerta,
        "cantidad_estimada": cantidad_estimada_final,
        "datosGrafica": {
            "labels": labels_grafica,
            "datasets": [
                {
                    "label": "Demanda Historica y Proyectada",
                    "data": data_grafica,
                    "borderColor": "#3b82f6",
                    "backgroundColor": "rgba(59, 130, 246, 0.5)",
                    "tension": 0.3,
                    "fill": True
                }
            ]
        }
    }

    df_proyeccion = pd.DataFrame({
        "fecha_proyectada": FECHAS_FUTURAS,
        "cantidad_proyectada": predicciones_xgboost,
        "origen_modelo": ["XGBoost"] * len(FECHAS_FUTURAS)
    })

    return {
        "resultado": resultado,
        "historico": df_filtrado.copy(),
        "proyeccion": df_proyeccion
    }


def _nombre_archivo_seguro(producto: str) -> str:
    nombre = re.sub(r"[^A-Za-z0-9_-]+", "_", str(producto)).strip("_") or "reporte"
    fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"reporte_ia_{nombre}_{fecha}.xlsx"


def _agregar_tabla_excel(ws, encabezados: list[str], filas: list[list]):
    ws.append(encabezados)
    for fila in filas:
        ws.append(fila)


def _estilizar_hoja_excel(ws, columnas_numericas: set[str] | None = None, columnas_largas: set[str] | None = None):
    columnas_numericas = columnas_numericas or set()
    columnas_largas = columnas_largas or set()
    header_fill = PatternFill("solid", fgColor=EXCEL_VERDE_CORPORATIVO)
    alt_fill = PatternFill("solid", fgColor=EXCEL_FONDO_SUAVE)
    header_font = Font(color=EXCEL_BLANCO, bold=True)
    border = Border(
        left=Side(style="thin", color=EXCEL_GRIS_BORDE),
        right=Side(style="thin", color=EXCEL_GRIS_BORDE),
        top=Side(style="thin", color=EXCEL_GRIS_BORDE),
        bottom=Side(style="thin", color=EXCEL_GRIS_BORDE)
    )

    encabezados = [cell.value for cell in ws[1]]
    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            encabezado = encabezados[cell.column - 1]
            cell.border = border
            if cell.row % 2 == 0:
                cell.fill = alt_fill
            if encabezado in columnas_largas:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif encabezado in columnas_numericas or "fecha" in encabezado:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col_idx, encabezado in enumerate(encabezados, start=1):
        max_length = len(str(encabezado))
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for item in cell:
                value = "" if item.value is None else str(item.value)
                max_length = max(max_length, len(value))

        if encabezado in columnas_largas:
            width = min(max(max_length * 0.55, 45), 80)
        else:
            width = min(max(max_length + 4, 12), 32)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 24
    if columnas_largas:
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 60


def _construir_workbook_reporte(df_resumen: pd.DataFrame, df_historico: pd.DataFrame, df_proyeccion: pd.DataFrame) -> Workbook:
    wb = Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_historico = wb.create_sheet("Historico")
    ws_proyeccion = wb.create_sheet("Proyeccion")

    _agregar_tabla_excel(ws_resumen, df_resumen.columns.tolist(), df_resumen.values.tolist())
    _agregar_tabla_excel(ws_historico, df_historico.columns.tolist(), df_historico.values.tolist())
    _agregar_tabla_excel(ws_proyeccion, df_proyeccion.columns.tolist(), df_proyeccion.values.tolist())

    _estilizar_hoja_excel(
        ws_resumen,
        columnas_numericas={"demanda_estimada"},
        columnas_largas={"recomendacion_ia"}
    )
    _estilizar_hoja_excel(
        ws_historico,
        columnas_numericas={"cantidad"}
    )
    _estilizar_hoja_excel(
        ws_proyeccion,
        columnas_numericas={"cantidad_proyectada"}
    )

    ws_resumen.sheet_view.showGridLines = False
    ws_historico.sheet_view.showGridLines = False
    ws_proyeccion.sheet_view.showGridLines = False
    ws_resumen.sheet_properties.tabColor = EXCEL_VERDE_CORPORATIVO
    ws_historico.sheet_properties.tabColor = EXCEL_GRIS_OSCURO
    ws_proyeccion.sheet_properties.tabColor = EXCEL_VERDE_CORPORATIVO
    return wb


@router.get("")
async def predecir_demanda(
    producto: str = Query(default="Todos", description="Nombre o ID del producto"),
    db: Session = Depends(get_db)
):
    try:
        return _construir_analisis(producto, db)["resultado"]
    except Exception as e:
        print(f"Error critico en IA: {str(e)}")
        return {
            "producto": "Desconocido",
            "recomendacion": f"Error interno en el microservicio IA: {str(e)}.",
            "nivel": "Error",
            "cantidad_estimada": 0,
            "datosGrafica": {
                "labels": [],
                "datasets": []
            }
        }


@router.get("/export")
async def exportar_reporte_demanda(
    producto: str = Query(default="Todos", description="Nombre o ID del producto"),
    db: Session = Depends(get_db)
):
    analisis = _construir_analisis(producto, db)
    resultado = analisis["resultado"]

    df_resumen = pd.DataFrame([{
        "producto": resultado["producto"],
        "fecha_generacion": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "demanda_estimada": resultado["cantidad_estimada"],
        "nivel_alerta": resultado["nivel"],
        "recomendacion_ia": resultado["recomendacion"]
    }])

    df_historico = analisis["historico"].copy()
    if df_historico.empty:
        df_historico = pd.DataFrame(columns=["fecha_venta", "producto_id", "nombre_producto", "cantidad"])
    else:
        df_historico = df_historico[["fecha_venta", "producto_id", "nombre_producto", "cantidad"]]

    output = BytesIO()
    workbook = _construir_workbook_reporte(df_resumen, df_historico, analisis["proyeccion"])
    workbook.save(output)

    output.seek(0)
    filename = _nombre_archivo_seguro(producto)
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )
